#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Sun Mar 19 21:08:11 2017

@author: mikewoodward
"""

from openpyxl import load_workbook

wb = load_workbook(
    filename='databreachesineurope-publicdata.xlsx')
print(f"Sheet names in workbook: {wb.sheetnames}")

# Master sheet
master_sheet = wb['Master']

# Excel starts row numbering from 1 and we want to start from 
# row 2, ignoring the header
for index in range(2, master_sheet.max_row + 1):
    leak_date = master_sheet[f'C{index}'].value
    organization = master_sheet[f'D{index}'].value
    records = master_sheet[f'F{index}'].value

    print(f"Leak date: {leak_date}, "
          f"Organization: {organization}, "
          f"Record count {records:,}")
